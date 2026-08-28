"""
Ablation over which same-domain reference audio candidate works best as the 1-shot/few-shot
example for ASR adaptation.

(Note: that using audio with the same ref_aud_tag as ref_aud/ref_text is best was already
established by the run_all_experiments.py experiment. This one asks which candidate to pick
from within those "same domain" candidates.)

Only the roads domain is run. Three criteria:
1. length  - candidates whose ref_text is 3/4/5 characters long (hardcoded in LENGTH_LEVEL_ENTITY_IDS)
2. phoneme - candidates whose ref_text has 6/9/12 unique phonemes (hardcoded in PHONEME_LEVEL_ENTITY_IDS)
3. concat  - few-shot built by concatenating 1/3/5 same-domain ref audios (chosen at random, once)

For criterion 3 (concat) the audio is really concatenated and written to
L2-KPNS/data/audio_ablation/, then fed straight into the existing 1-shot
run_asr_batch(ref_audio=..., reference_text=...) pipeline (the model pipeline code is untouched).

python -m src.run_ablation_experiments                        # all of ABLATION_SPEAKERS
python -m src.run_ablation_experiments --speakers P001,P002   # a subset (summary merge is skipped)
python -m src.run_ablation_experiments --merge-only           # merge per-speaker summaries, no inference

To run several jobs at once, split the speakers across --speakers so they do not overlap.
Both the result jsonl and the summary csv are separated per speaker, so as long as the speakers
do not overlap the jobs cannot collide.
"""

from __future__ import annotations

import argparse
import csv
import importlib.util
import json
import multiprocessing as mp
import os
import queue
import random
import re
import shutil
import statistics
import sys
from collections import defaultdict
from typing import List, Optional, Tuple

import numpy as np
import openpyxl
import soundfile as sf
from openpyxl.styles import Font
from tqdm import tqdm

import compute_metrics

# infer_qwen2.5_omni_batch.py has a "." in its filename, so it cannot be pulled in with a
# plain import statement -- load it by path via importlib. dataclass consults sys.modules when
# resolving forward refs, so the module must be registered there before exec_module.
_spec = importlib.util.spec_from_file_location(
    "infer_qwen2_5_omni_batch",
    os.path.join(os.path.dirname(os.path.abspath(__file__)), "infer_qwen2.5_omni_batch.py"),
)
infer_mod = importlib.util.module_from_spec(_spec)
sys.modules["infer_qwen2_5_omni_batch"] = infer_mod
_spec.loader.exec_module(infer_mod)

AUDIO_ROOT = "L2-KPNS/data/audio"
ABLATION_AUDIO_ROOT = "tmp/audio_ablation"  # scratch space for the concatenated (shot3/5) audio; deleted when the run finishes
DOMAINS = ["roads"]
DOMAIN_TAG = {"roads": "rd", "content": "ct", "restaurants": "rt", "stations": "st"}
WAKEWORD_CSV = "L2-KPNS/metadata/recording_targets/wakeword_candidates.csv"

BATCH_SIZE = 8
# Criteria 1/2: entity_ids picked by hand from the measured roads-domain stats
# (ref_text_domain_stats.xlsx, roads sheet).
# length: 서달로 (3 chars) / 영선대로 (4 chars) / 후평도원로 (5 chars)
LENGTH_LEVEL_ENTITY_IDS = ["pr0021_rd", "pr0022_rd", "pr0023_rd"]
# phoneme: 서달로 (6) / 영선대로 (9) / 향소동막골길 (12)
PHONEME_LEVEL_ENTITY_IDS = ["pr0021_rd", "pr0022_rd", "pr0028_rd"]
SHOT_COUNTS = [1, 3, 5]  # criterion 3: number of clips concatenated
NUM_ABLATION_SPEAKERS = 14  # sanity check
RANDOM_SEED = 42
SILENCE_GAP_SEC = 0.3    # silence inserted between clips when concatenating

ABLATION_SPEAKERS = [
    'P001', 'P002', 'P003', 'P004', 'P005',
    'P009', 'P010', 'P011', 'P012',
    'P014', 'P015',
    'P017',
    'P019', 'P020',
]

RESULTS_ROOT = "L2-KPNS-jsonl/results_ablation"
SUMMARY_CSV = f"{RESULTS_ROOT}/summary_ablation.csv"
SUMMARY_XLSX = f"{RESULTS_ROOT}/summary_ablation.xlsx"
SUMMARY_FIELDS = [
    "speaker", "domain", "criterion", "level", "num_shots",
    "entity_ids", "ref_texts", "ref_text_lengths", "unique_phoneme_counts",
    "ref_audio_file", "num_rows",
    "WER", "PED", "ASR-Hit", "Hit@1",
    "jsonl_path",
]


def list_speakers() -> List[str]:
    return sorted(
        d for d in os.listdir(AUDIO_ROOT)
        if os.path.isdir(os.path.join(AUDIO_ROOT, d))
    )


def load_domain_candidates() -> dict:
    """domain -> [{"entity_id":..., "korean":...}, ...], read from wakeword_candidates.csv
    (the domain tag is taken from the entity_id suffix)."""
    by_tag = {tag: [] for tag in DOMAIN_TAG.values()}
    with open(WAKEWORD_CSV, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            tag = row["entity_id"].rsplit("_", 1)[-1]
            if tag in by_tag:
                by_tag[tag].append({
                    "entity_id": row["entity_id"],
                    "korean": row["korean"].strip().rstrip("."),
                })
    return {domain: by_tag[tag] for domain, tag in DOMAIN_TAG.items()}


def ref_audio_path(entity_id: str, speaker: str) -> str:
    return f"{AUDIO_ROOT}/{speaker}/{entity_id}_{speaker}.wav"


def valid_candidates_for_speaker(candidates: List[dict], speaker: str) -> List[dict]:
    valid = []
    for c in candidates:
        path = ref_audio_path(c["entity_id"], speaker)
        if os.path.exists(path):
            valid.append(c)
        else:
            print(f"  [error] {path} 없음, 후보 {c['entity_id']} 제외")
    return valid


def annotate_metrics(candidates: List[dict]) -> None:
    """Fill in the ref_text length and unique phoneme count on each candidate dict, in place."""
    for c in candidates:
        c["length"] = len(c["korean"])
        c["unique_phoneme_count"] = len(set(infer_mod.epi.transliterate(c["korean"])))


def pick_by_entity_id(candidates: List[dict], entity_id: str) -> dict:
    for c in candidates:
        if c["entity_id"] == entity_id:
            return c
    raise ValueError(f"entity_id={entity_id} 인 후보가 없음")


def build_entity_dict_ipa(domain: str) -> dict:
    lexicon_path = f"L2-KPNS/metadata/recording_targets/{domain}_200.csv"
    with open(lexicon_path, encoding="utf-8-sig") as f:
        entity_dict = {row["entity_id"]: row["korean"] for row in csv.DictReader(f)}
    entity_names = list(dict.fromkeys(entity_dict.values()))  # dedupe, preserve order
    return {e: infer_mod.epi.transliterate(e) for e in entity_names}


def build_ref_audio(chosen: List[dict], speaker: str, save_path: str) -> Tuple[np.ndarray, str, str]:
    """Concatenate the audio of the chosen candidates in order with a silence gap between them,
    concatenate their ref_texts too, and return an (audio, text) pair that can be fed straight
    into the existing 1-shot pipeline (run_asr_batch).
    With two or more candidates the concatenated audio is written to save_path as wav; with a
    single candidate the original file is reused as-is (so audio_ablation does not end up
    holding a duplicate of the original)."""
    if len(chosen) == 1:
        c = chosen[0]
        original_path = ref_audio_path(c["entity_id"], speaker)
        return infer_mod.load_audio_16k(original_path), c["korean"], original_path

    gap = np.zeros(int(SILENCE_GAP_SEC * infer_mod.TARGET_SR), dtype=np.float32)
    clips = []
    for i, c in enumerate(chosen):
        if i > 0:
            clips.append(gap)
        clips.append(infer_mod.load_audio_16k(ref_audio_path(c["entity_id"], speaker)))
    concatenated = np.concatenate(clips)
    ref_text = " ".join(c["korean"] for c in chosen)

    os.makedirs(os.path.dirname(save_path) or ".", exist_ok=True)
    sf.write(save_path, concatenated, infer_mod.TARGET_SR)
    return concatenated, ref_text, save_path


def run_one_ablation_config(
    asr: "infer_mod.Qwen25OmniASRPipeline",
    speaker: str,
    domain: str,
    data: List[dict],
    entity_dict_ipa: dict,
    ref_audio: np.ndarray,
    ref_text: str,
    ref_audio_file: str,
    criterion: str,
    level: str,
    chosen: List[dict],
    gpu_id: int,
) -> Optional[str]:
    """Run one combination in batches of BATCH_SIZE and return the path of the result jsonl.

    If any audio file is missing, the whole combination is skipped (returns None)."""
    results = []
    pbar = tqdm(total=len(data), desc=f"[GPU {gpu_id}] {speaker}/{domain}/{criterion}/{level}", leave=False)
    for batch_rows in infer_mod.chunked(data, BATCH_SIZE):
        batch_audio_paths = [f"{AUDIO_ROOT}/{speaker}/" + row["audio_path"] for row in batch_rows]
        missing = next((p for p in batch_audio_paths if not os.path.exists(p)), None)
        if missing is not None:
            pbar.close()
            print(f"  [error] {missing} 없음, 이 조합({speaker}/{domain}/{criterion}/{level}) 전체 skip")
            return None
        batch_audio = [infer_mod.load_audio_16k(p) for p in batch_audio_paths]

        batch_out = asr.run_asr_batch(
            audios=batch_audio, ref_audio=ref_audio, reference_text=ref_text,
        )

        for row, out in zip(batch_rows, batch_out):
            entity_list_text = [e[0] for e in infer_mod.retrieve_top_k(
                utterance=out, entity_ipa=entity_dict_ipa, k=10,
            )]
            result = dict(row)
            result["asr_result"] = out
            result["retr_entities"] = entity_list_text
            results.append(result)
        pbar.update(len(batch_rows))
    pbar.close()

    save_jsonl = f"{RESULTS_ROOT}/{speaker}/{domain}_{criterion}_{level}.jsonl"
    os.makedirs(os.path.dirname(save_jsonl) or ".", exist_ok=True)
    with open(save_jsonl, "w", encoding="utf-8") as f:
        for row in results:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")

    metrics = compute_metrics.compute_metrics(results)
    write_summary_row(
        speaker, domain, criterion, level, len(chosen), chosen,
        ref_audio_file, len(results), metrics, save_jsonl,
    )
    return save_jsonl


def write_summary_row(
    speaker, domain, criterion, level, num_shots, chosen, ref_audio_file, num_rows, metrics, jsonl_path,
):
    row = {
        "speaker": speaker,
        "domain": domain,
        "criterion": criterion,
        "level": level,
        "num_shots": num_shots,
        "entity_ids": "|".join(c["entity_id"] for c in chosen),
        "ref_texts": "|".join(c["korean"] for c in chosen),
        "ref_text_lengths": "|".join(str(c["length"]) for c in chosen),
        "unique_phoneme_counts": "|".join(str(c["unique_phoneme_count"]) for c in chosen),
        "ref_audio_file": ref_audio_file,
        "num_rows": num_rows,
        "WER": metrics["wer"][0],
        "PED": metrics["phoneme_editdistance"][0],
        "ASR-Hit": metrics["recall@asr"][0],
        "Hit@1": metrics["recall@1"][0],
        "jsonl_path": jsonl_path,
    }
    with open(speaker_summary_csv(speaker), "a", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=SUMMARY_FIELDS)
        writer.writerow(row)


def speaker_summary_csv(speaker: str) -> str:
    """Summary rows go into a per-speaker file. Whether there are several GPU workers or several
    jobs split by --speakers, no two processes ever touch the same file, so rows can never be
    interleaved or duplicated."""
    return f"{RESULTS_ROOT}/{speaker}/summary.csv"


def init_speaker_summary_csv(speaker: str) -> None:
    """Create the speaker summary file from scratch (re-running a speaker discards its old rows)."""
    path = speaker_summary_csv(speaker)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="") as f:
        csv.DictWriter(f, fieldnames=SUMMARY_FIELDS).writeheader()


def merge_speaker_summaries(speakers: List[str]) -> None:
    """Collect the per-speaker summary files and rewrite the summary csv (rewritten every time
    rather than appended, so the same combination can never appear twice)."""
    rows = []
    missing = []
    for speaker in speakers:
        path = speaker_summary_csv(speaker)
        if not os.path.exists(path):
            missing.append(speaker)
            continue
        with open(path, encoding="utf-8", newline="") as f:
            rows.extend(csv.DictReader(f))
    if missing:
        print(f"  [warn] 요약 파일이 없어 제외된 speaker: {missing}")

    speaker_order = {s: i for i, s in enumerate(speakers)}
    criterion_order = {c: i for i, c in enumerate(("length", "phoneme", "concat"))}
    rows.sort(key=lambda r: (
        speaker_order.get(r["speaker"], len(speaker_order)),
        r["domain"],
        criterion_order.get(r["criterion"], len(criterion_order)),
        _level_sort_key(r["level"]),
    ))

    os.makedirs(os.path.dirname(SUMMARY_CSV) or ".", exist_ok=True)
    with open(SUMMARY_CSV, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=SUMMARY_FIELDS)
        writer.writeheader()
        writer.writerows(rows)
    print(f"Merged {len(rows)} rows from {len(speakers) - len(missing)} speaker(s) into {SUMMARY_CSV}")


METRIC_FIELDS = ["WER", "PED", "ASR-Hit", "Hit@1"]


def _level_sort_key(level: str):
    """Pull the leading number out of a level string (len3_..., phon6_..., shot1) and use it as
    an ascending sort key."""
    m = re.match(r"[a-z]+(\d+)", level)
    return int(m.group(1)) if m else level


def write_level_average_sheet(wb: "openpyxl.Workbook", rows: List[dict]) -> None:
    """A sheet averaging over all speakers per criterion/level, to see the speaker-independent trend."""
    groups: dict = defaultdict(list)
    for row in rows:
        groups[(row["criterion"], row["level"])].append(row)

    ws = wb.create_sheet(title="avg_by_level")
    ws.append(["criterion", "level", "num_speakers", *METRIC_FIELDS])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for criterion in ("length", "phoneme", "concat"):
        keys = sorted(
            (k for k in groups if k[0] == criterion),
            key=lambda k: _level_sort_key(k[1]),
        )
        for key in keys:
            group_rows = groups[key]
            means = [
                round(statistics.mean(float(r[field]) for r in group_rows), 2)
                for field in METRIC_FIELDS
            ]
            ws.append([key[0], key[1], len(group_rows), *means])


def convert_summary_csv_to_xlsx() -> None:
    """After every process has finished, convert the final summary_ablation.csv to
    summary_ablation.xlsx, with a raw sheet (all original rows) and an avg_by_level sheet
    (speaker averages per criterion/level)."""
    wb = openpyxl.Workbook()
    raw_ws = wb.active
    raw_ws.title = "raw"
    with open(SUMMARY_CSV, encoding="utf-8", newline="") as f:
        for row in csv.reader(f):
            raw_ws.append(row)

    with open(SUMMARY_CSV, encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    write_level_average_sheet(wb, rows)

    wb.save(SUMMARY_XLSX)


def run_for_speaker(
    asr: "infer_mod.Qwen25OmniASRPipeline",
    domain_candidates: dict,
    speaker: str,
    gpu_id: int,
) -> None:
    init_speaker_summary_csv(speaker)
    rng = random.Random(RANDOM_SEED)  # fixed seed so every speaker uses the same concat sample

    for domain in DOMAINS:
        data_jsonl = f"L2-KPNS-jsonl/{speaker}/{domain}_{speaker}.jsonl"
        data = infer_mod.load_jsonl(data_jsonl)
        entity_dict_ipa = build_entity_dict_ipa(domain)

        candidates = valid_candidates_for_speaker(domain_candidates[domain], speaker)
        if not candidates:
            print(f"[GPU {gpu_id}] {speaker}/{domain}: 사용 가능한 ref 후보 없음, domain skip")
            continue
        annotate_metrics(candidates)

        # ---- Criterion 1: ref_text length (hardcoded in LENGTH_LEVEL_ENTITY_IDS) ----
        for entity_id in LENGTH_LEVEL_ENTITY_IDS:
            c = pick_by_entity_id(candidates, entity_id)
            level = f"len{c['length']}_{c['entity_id']}"
            save_path = f"{ABLATION_AUDIO_ROOT}/{speaker}/{domain}_length_{level}.wav"
            ref_audio, ref_text, ref_audio_file = build_ref_audio([c], speaker, save_path)
            run_one_ablation_config(
                asr, speaker, domain, data, entity_dict_ipa,
                ref_audio=ref_audio, ref_text=ref_text, ref_audio_file=ref_audio_file,
                criterion="length", level=level, chosen=[c], gpu_id=gpu_id,
            )

        # ---- Criterion 2: unique phoneme count (hardcoded in PHONEME_LEVEL_ENTITY_IDS) ----
        for entity_id in PHONEME_LEVEL_ENTITY_IDS:
            c = pick_by_entity_id(candidates, entity_id)
            level = f"phon{c['unique_phoneme_count']}_{c['entity_id']}"
            save_path = f"{ABLATION_AUDIO_ROOT}/{speaker}/{domain}_phoneme_{level}.wav"
            ref_audio, ref_text, ref_audio_file = build_ref_audio([c], speaker, save_path)
            run_one_ablation_config(
                asr, speaker, domain, data, entity_dict_ipa,
                ref_audio=ref_audio, ref_text=ref_text, ref_audio_file=ref_audio_file,
                criterion="phoneme", level=level, chosen=[c], gpu_id=gpu_id,
            )

        # ---- Criterion 3: number of concatenated clips (1/3/5, few-shot, chosen at random once) ----
        for num_shots in SHOT_COUNTS:
            if num_shots > len(candidates):
                print(f"[GPU {gpu_id}] {speaker}/{domain}: 후보가 {len(candidates)}개뿐이라 "
                      f"{num_shots}-shot skip")
                continue
            chosen = rng.sample(candidates, num_shots)
            level = f"shot{num_shots}"
            save_path = f"{ABLATION_AUDIO_ROOT}/{speaker}/{domain}_concat_{level}.wav"
            ref_audio, ref_text, ref_audio_file = build_ref_audio(chosen, speaker, save_path)
            run_one_ablation_config(
                asr, speaker, domain, data, entity_dict_ipa,
                ref_audio=ref_audio, ref_text=ref_text, ref_audio_file=ref_audio_file,
                criterion="concat", level=level, chosen=chosen, gpu_id=gpu_id,
            )


def gpu_worker(gpu_id: int, speaker_queue: "mp.Queue[str]", domain_candidates: dict) -> None:
    """One worker per GPU. Speakers are pulled off the queue one at a time and run sequentially,
    so even with more speakers than GPUs each GPU keeps exactly one process running."""
    asr = infer_mod.Qwen25OmniASRPipeline(infer_mod.ASRConfig(device=f"cuda:{gpu_id}"))

    while True:
        try:
            speaker = speaker_queue.get_nowait()
        except queue.Empty:
            break
        run_for_speaker(asr, domain_candidates, speaker, gpu_id)


def finalize_summary() -> None:
    """Merge the per-speaker summaries into the summary csv and convert it to xlsx
    (raw + avg_by_level)."""
    merge_speaker_summaries(ABLATION_SPEAKERS)
    convert_summary_csv_to_xlsx()
    print(f"Saved summary to {SUMMARY_XLSX}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--speakers", default=None,
        help="Comma-separated speaker list. Defaults to all of ABLATION_SPEAKERS. "
             "When running several jobs at once, split them so they do not overlap.",
    )
    parser.add_argument(
        "--merge-only", action="store_true",
        help="Skip inference and just merge the per-speaker summary files into the summary csv/xlsx.",
    )
    args = parser.parse_args()

    if args.merge_only:
        finalize_summary()
        sys.exit(0)

    mp.set_start_method("spawn", force=True)

    ablation_speakers = (
        [s.strip() for s in args.speakers.split(",")] if args.speakers else list(ABLATION_SPEAKERS)
    )
    domain_candidates = load_domain_candidates()

    num_gpus = infer_mod.torch.cuda.device_count() or 1
    print(f"Detected {num_gpus} GPU(s); processing {len(ablation_speakers)} "
          f"speaker(s) {ablation_speakers} via a queue.")

    speaker_queue: "mp.Queue[str]" = mp.Queue()
    for speaker in ablation_speakers:
        speaker_queue.put(speaker)

    procs = []
    for gpu_id in range(num_gpus):
        p = mp.Process(
            target=gpu_worker, args=(gpu_id, speaker_queue, domain_candidates), name=f"gpu-{gpu_id}",
        )
        p.start()
        procs.append(p)
    for p in procs:
        p.join()

    if args.speakers:
        # Another job may still be running the remaining speakers, so do not merge here.
        print("--speakers 로 일부만 실행했으므로 요약 병합은 생략합니다. "
              "모든 잡이 끝난 뒤 `python run_ablation_experiments.py --merge-only` 를 실행하세요.")
    else:
        finalize_summary()

    if os.path.isdir(f"{ABLATION_AUDIO_ROOT}"):
        # The concat scratch audio lives in per-speaker subdirectories, so only remove the
        # speakers this job ran (another job on the same node may still be using its own).
        for speaker in ablation_speakers:
            speaker_dir = f"{ABLATION_AUDIO_ROOT}/{speaker}"
            if os.path.isdir(speaker_dir):
                shutil.rmtree(speaker_dir)
        print(f"Removed temp audio for {len(ablation_speakers)} speaker(s) under {ABLATION_AUDIO_ROOT}")
